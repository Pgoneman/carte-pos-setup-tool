"""
CARTE POS Setup Tool
Upload a menu Excel or paste a BentoBox / Google Maps URL → instant CARTE POS setup file.
"""

import os
import re
import json
from datetime import datetime
from urllib.parse import urlparse
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import requests as http_req
from bs4 import BeautifulSoup

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

if os.environ.get('VERCEL'):
    UPLOAD_DIR = '/tmp/uploads'
else:
    UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
CARTE_API = 'http://39.125.162.234:8080'

# ─── Helpers ───────────────────────────────────────────────────────────────

def parse_price(raw):
    if raw is None: return 0.0
    t = re.sub(r'[^\d.]', '', str(raw).strip())
    try: return round(float(t), 2)
    except: return 0.0

def clean(v):
    s = str(v or '').strip()
    return '' if s.lower() in ('', 'none', 'nan') else s


def is_bentobox_url(url):
    """Check if a URL is a BentoBox restaurant website."""
    return bool(re.search(r'\.getbento\.com', url, re.IGNORECASE))


def extract_bentobox_base(url):
    """Extract the base BentoBox URL (https://{slug}.getbento.com)."""
    parsed = urlparse(url if '://' in url else 'https://' + url)
    host = parsed.hostname or ''
    m = re.match(r'([\w-]+\.getbento\.com)', host)
    if m:
        return f'https://{m.group(1)}'
    return None


# ─── BentoBox API Integration ────────────────────────────────────────────

# Allergen boolean fields on BentoBox products → grouped for output
ALLERGEN_FIELDS = [
    'fish', 'shellfish', 'dairy', 'egg', 'gluten', 'wheat',
    'peanut', 'tree_nut', 'soy', 'sesame', 'corn',
]
DIETARY_FIELDS = [
    'vegan', 'vegetarian', 'gluten_free', 'dairy_free',
    'halal', 'kosher', 'keto', 'paleo', 'plant_based',
    'organic', 'raw', 'spicy', 'mild', 'medium', 'hot',
]


def _extract_flags(product, field_list):
    """Extract truthy boolean flags from a BentoBox product dict."""
    return [f for f in field_list if product.get(f)]


def _find_menu_id(location):
    """Find a valid menu_id from a BentoBox location's fulfillment_options.
    Checks pickup, delivery, and dine_in in order."""
    opts = location.get('fulfillment_options', {})
    for ftype in ('pickup', 'delivery', 'dine_in'):
        foption = opts.get(ftype, {})
        datetimes = foption.get('datetimes_with_ranges', [])
        if datetimes:
            hour_ranges = datetimes[0].get('hour_ranges', [])
            if hour_ranges:
                mid = hour_ranges[0].get('menu_id')
                if mid:
                    return mid
    return None


def _extract_hours_from_location(location):
    """Extract structured hours from a BentoBox location."""
    hours = {}
    for htype in ('pickup', 'delivery', 'dine_in'):
        key = f'open_{htype}_hours'
        hdata = location.get(key, {})
        weekday = hdata.get('weekday_hours', {})
        if weekday:
            hours[htype] = weekday
    return hours


def _extract_business_name_from_html(base_url, session):
    """Try to extract the business name from JSON-LD on the main page."""
    try:
        resp = session.get(base_url, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                data = json.loads(script.string)
                if isinstance(data, dict):
                    if data.get('@type') == 'Organization' and data.get('name'):
                        return data['name']
                    if data.get('name'):
                        return data['name']
                elif isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict) and item.get('name'):
                            return item['name']
            except (json.JSONDecodeError, TypeError):
                continue
        # Fallback: parse <title>
        title_tag = soup.find('title')
        if title_tag and title_tag.string:
            name = title_tag.string.strip()
            # Remove common suffixes
            name = re.sub(r'\s*[|–-]\s*(Online Ordering|Menu|Restaurant|Home).*$', '', name, flags=re.IGNORECASE)
            return name.strip()
    except Exception:
        pass
    # Last fallback: derive from subdomain
    parsed = urlparse(base_url)
    subdomain = (parsed.hostname or '').split('.')[0]
    return subdomain.replace('-', ' ').title()


def fetch_bentobox_data(url):
    """Fetch full menu data from a BentoBox restaurant website.

    Returns a dict with business_name, locations (each with address, hours, menu).
    """
    base_url = extract_bentobox_base(url)
    if not base_url:
        raise ValueError(f'Not a valid BentoBox URL: {url}')

    s = http_req.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
    })

    # Step 1: Get session + CSRF from ordering page
    try:
        s.get(f'{base_url}/online-ordering/', timeout=15)
    except Exception as e:
        raise ValueError(f'Could not reach BentoBox site: {e}')

    csrf = s.cookies.get('csrftoken', '')
    if csrf:
        s.headers['X-CSRFToken'] = csrf

    # Get business name from main page JSON-LD
    business_name = _extract_business_name_from_html(base_url, s)

    # Step 2: Get locations
    try:
        resp = s.get(f'{base_url}/api/online_ordering_location_public/', timeout=15)
        resp.raise_for_status()
        raw_locations = resp.json()
    except Exception as e:
        raise ValueError(f'Failed to fetch BentoBox locations: {e}')

    if not raw_locations:
        raise ValueError('No locations found on this BentoBox site')

    locations = []
    for loc in raw_locations:
        loc_data = loc.get('location', {})
        loc_id = loc.get('id')

        # Build address
        address = {
            'street': loc_data.get('street', ''),
            'city': loc_data.get('city', ''),
            'state': loc_data.get('state', ''),
            'zip': loc_data.get('postal_code', ''),
        }

        # Find menu_id
        menu_id = _find_menu_id(loc)

        # Extract hours
        hours = _extract_hours_from_location(loc)

        location_entry = {
            'id': loc_id,
            'name': loc_data.get('name', ''),
            'slug': loc_data.get('slug', ''),
            'address': address,
            'phone': loc_data.get('phone_number', ''),
            'lat': str(loc_data.get('lat', '')),
            'lng': str(loc_data.get('lng', '')),
            'hours': hours,
            'menu_id': menu_id,
            'menu': {'categories': []},
        }

        # Step 3 + 4: Init ordering and get menu for this location
        if menu_id and loc_id:
            try:
                # Re-fetch CSRF for safety
                s.get(f'{base_url}/online-ordering/', timeout=10)
                csrf = s.cookies.get('csrftoken', '')
                if csrf:
                    s.headers['X-CSRFToken'] = csrf

                # Init ordering session
                s.put(f'{base_url}/api/online_ordering/initial_data/',
                      json={'location_id': loc_id, 'commerce_type': 'online_ordering'},
                      timeout=15)

                # Get menu
                menu_resp = s.get(f'{base_url}/api/online_ordering/menu/{menu_id}/', timeout=20)
                menu_resp.raise_for_status()
                menu_data = menu_resp.json()

                categories = []
                for section in menu_data.get('sections', []):
                    cat = {
                        'name': section.get('name', 'Uncategorized'),
                        'description': section.get('description', ''),
                        'items': [],
                    }
                    for item_wrapper in section.get('items', []):
                        product = item_wrapper.get('product', {})
                        # Image URL
                        images = product.get('images', [])
                        image_url = images[0].get('url', '') if images else ''

                        # Allergens and dietary flags
                        allergens = _extract_flags(product, ALLERGEN_FIELDS)
                        dietary = _extract_flags(product, DIETARY_FIELDS)

                        # Variants
                        variants = []
                        for v in product.get('variants', []):
                            variants.append({
                                'name': v.get('name', ''),
                                'price': parse_price(v.get('price')),
                                'weight': v.get('weight', ''),
                                'calories': v.get('calories', ''),
                            })

                        # Default price
                        price = parse_price(product.get('default_price'))
                        if not price and variants:
                            price = variants[0].get('price', 0)

                        cat['items'].append({
                            'name': product.get('name', ''),
                            'price': price,
                            'description': product.get('description', ''),
                            'image_url': image_url,
                            'allergens': allergens,
                            'dietary': dietary,
                            'variants': variants,
                        })
                    categories.append(cat)

                location_entry['menu'] = {'categories': categories}
            except Exception:
                # Menu fetch failed for this location - continue with empty menu
                pass

        locations.append(location_entry)

    return {
        'success': True,
        'source': 'bentobox',
        'business_name': business_name,
        'locations': locations,
    }


def bentobox_to_parsed(bb_data):
    """Convert BentoBox API data to the internal parsed format used by Excel generation."""
    parsed = {
        'stores': [],
        'hours': [],
        'menus': {},
        'options': [],
        'compare': [],
        'sheets_found': ['BentoBox Import'],
    }

    biz_name = bb_data.get('business_name', 'Store')

    for loc in bb_data.get('locations', []):
        addr = loc.get('address', {})
        store_name = loc.get('name') or biz_name
        full_address = ', '.join(filter(None, [
            addr.get('street', ''), addr.get('city', ''),
            addr.get('state', ''), addr.get('zip', ''),
        ]))

        parsed['stores'].append({
            'business_name': biz_name,
            'name': store_name,
            'address': full_address,
            'phone': loc.get('phone', ''),
            'lat': loc.get('lat', ''),
            'lng': loc.get('lng', ''),
            'street': addr.get('street', ''),
            'city': addr.get('city', ''),
            'state': addr.get('state', ''),
            'zip': addr.get('zip', ''),
        })

        # Hours
        day_names = {
            '0': 'Mon', '1': 'Tue', '2': 'Wed', '3': 'Thu',
            '4': 'Fri', '5': 'Sat', '6': 'Sun',
        }
        for htype, weekday_hours in loc.get('hours', {}).items():
            label = htype.replace('_', ' ').title()
            if isinstance(weekday_hours, dict):
                for day_num, ranges in weekday_hours.items():
                    day_label = day_names.get(str(day_num), str(day_num))
                    if isinstance(ranges, list):
                        for r in ranges:
                            if isinstance(r, dict):
                                parsed['hours'].append({
                                    'store': store_name,
                                    'type': label,
                                    'days': day_label,
                                    'open': r.get('open', ''),
                                    'close': r.get('close', ''),
                                })
                            elif isinstance(r, (list, tuple)) and len(r) >= 2:
                                parsed['hours'].append({
                                    'store': store_name,
                                    'type': label,
                                    'days': day_label,
                                    'open': str(r[0]),
                                    'close': str(r[1]),
                                })

        # Menu items
        menu_items = []
        for cat in loc.get('menu', {}).get('categories', []):
            for item in cat.get('items', []):
                allergens = item.get('allergens', [])
                dietary = item.get('dietary', [])
                allergen_str = ', '.join(allergens + dietary)
                menu_items.append({
                    'category': cat.get('name', ''),
                    'name': item.get('name', ''),
                    'price': item.get('price', 0),
                    'description': item.get('description', ''),
                    'image_url': item.get('image_url', ''),
                    'allergens': allergens,
                    'dietary': dietary,
                    'allergen_info': allergen_str,
                    'option_groups': '',
                    'options_text': '',
                })
        parsed['menus'][store_name] = menu_items

    return parsed


# ─── Smart Excel Parser ───────────────────────────────────────────────────
# Handles various formats: Sushi Kudasai multi-sheet, simple menu lists, etc.

def parse_uploaded_excel(filepath):
    """Parse any menu Excel and extract store info, hours, menus, options."""
    wb = load_workbook(filepath, data_only=True)
    result = {
        'stores': [],       # list of {name, address, phone, lat, lng, ...}
        'hours': [],        # list of {store, type, days, open, close}
        'menus': {},        # store_name -> [{category, name, price, description, options_text}]
        'options': [],      # [{store, item, group, option, price, type}]
        'compare': [],
        'sheets_found': wb.sheetnames,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        name_lower = sheet_name.lower()

        # ── Priority 1: Match by sheet name first ──
        if any(k in name_lower for k in ['store info', 'store_info', '매장', '가게']):
            result['stores'] = _parse_store_info(rows)
        elif any(k in name_lower for k in ['hour', '시간', '영업', 'schedule']):
            result['hours'] = _parse_hours(rows)
        elif any(k in name_lower for k in ['option', '옵션']) and 'group' not in name_lower:
            result['options'] = _parse_options(rows)
        elif any(k in name_lower for k in ['compare', '비교']):
            result['compare'] = _parse_compare(rows)
        elif any(k in name_lower for k in ['menu', '메뉴']):
            store_label = _extract_store_label(sheet_name)
            result['menus'][store_label] = _parse_menu_sheet(rows)
        # ── Priority 2: Content-based detection ──
        elif _is_store_info_sheet(rows):
            result['stores'] = _parse_store_info(rows)
        elif _is_hours_sheet(rows, name_lower):
            result['hours'] = _parse_hours(rows)
        elif _is_options_sheet(rows, name_lower):
            result['options'] = _parse_options(rows)
        elif _is_menu_sheet(rows, name_lower):
            store_label = _extract_store_label(sheet_name)
            result['menus'][store_label] = _parse_menu_sheet(rows)

    # If no menu was found under a store label, try first sheet
    if not result['menus']:
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        result['menus']['Default'] = _parse_menu_sheet(rows)

    wb.close()
    return result


def _is_store_info_sheet(rows):
    """Detect store info by looking for address/phone keywords."""
    text = ' '.join(clean(c) for row in rows[:10] for c in (row or []))
    keywords = ['address', '주소', 'phone', '전화', 'location', 'lat', '위도']
    return sum(1 for k in keywords if k.lower() in text.lower()) >= 2

def _is_hours_sheet(rows, name):
    if 'hour' in name or '시간' in name or '영업' in name: return True
    text = ' '.join(clean(c) for row in rows[:5] for c in (row or []))
    return any(k in text.lower() for k in ['mon', 'tue', 'pickup', 'delivery', 'dine'])

def _is_menu_sheet(rows, name):
    if 'menu' in name or '메뉴' in name: return True
    # Check if rows have price-like data
    price_count = 0
    for row in rows[1:20]:
        for c in (row or []):
            if c and '$' in str(c): price_count += 1
    return price_count >= 3

def _is_options_sheet(rows, name):
    if 'option' in name or '옵션' in name: return True
    header = ' '.join(clean(c) for c in (rows[0] or []))
    return 'option' in header.lower() or '옵션' in header

def _extract_store_label(sheet_name):
    """Extract store name from sheet title like 'Menu - US Bank Centre'."""
    m = re.search(r'[-–]\s*(.+)', sheet_name)
    return m.group(1).strip() if m else sheet_name.replace('Menu', '').strip() or 'Default'


def _parse_store_info(rows):
    """Parse store info sheet. Handles multi-location column layout."""
    stores = []
    # Find how many location columns (skip first label column)
    header_row = None
    for i, row in enumerate(rows[:10]):
        vals = [clean(c) for c in (row or [])]
        non_empty = [v for v in vals[1:] if v]
        # Header row: has 2+ values in B,C,D columns AND these look like location names (not long descriptions)
        if len(non_empty) >= 2 and all(len(v) < 60 for v in non_empty):
            # Check the NEXT rows also have 2+ values (it's a real multi-column layout)
            if i + 1 < len(rows):
                next_vals = [clean(c) for c in (rows[i+1] or [])]
                next_non_empty = [v for v in next_vals[1:] if v]
                if len(next_non_empty) >= 2:
                    header_row = i
                    break

    if header_row is None:
        # Single store - look for key-value pairs
        store = {}
        for row in rows:
            vals = [clean(c) for c in (row or [])]
            if len(vals) >= 2 and vals[0]:
                key = vals[0].lower()
                if any(k in key for k in ['name', '이름', '상호']): store['name'] = vals[1]
                elif any(k in key for k in ['address', '주소']): store['address'] = vals[1]
                elif any(k in key for k in ['phone', '전화']): store['phone'] = vals[1]
                elif any(k in key for k in ['lat', '위도']): store['lat'] = vals[1]
                elif any(k in key for k in ['lng', 'lon', '경도']): store['lng'] = vals[1]
        if store: stores.append(store)
        return stores

    # Multi-location: columns B, C, D... are different locations
    location_names = [clean(c) for c in (rows[header_row] or [])[1:] if clean(c)]
    for ln in location_names:
        stores.append({'name': ln})

    # Parse key-value rows
    for row in rows[header_row+1:]:
        vals = [clean(c) for c in (row or [])]
        if not vals or not vals[0]: continue
        key = vals[0].lower()
        for idx, store in enumerate(stores):
            val = vals[idx+1] if idx+1 < len(vals) else ''
            if not val: continue
            if any(k in key for k in ['address', '주소']): store['address'] = val
            elif any(k in key for k in ['phone', '전화']): store['phone'] = val
            elif any(k in key for k in ['lat', '위도']): store['lat'] = val
            elif any(k in key for k in ['lng', 'lon', '경도']): store['lng'] = val
            elif any(k in key for k in ['온라인', 'online', '배달']): store['online_order'] = val
            elif any(k in key for k in ['dine', '매장', '식사']): store['dine_in'] = val
            elif any(k in key for k in ['메뉴 수', 'menu count']): store['menu_count'] = val
            elif any(k in key for k in ['홈페이지', 'website', 'homepage']): store['website'] = val
            elif any(k in key for k in ['온라인주문', 'order', 'bento']): store['online_url'] = val

    # Also grab the business name from first row
    first_text = clean(rows[0][0]) if rows[0] else ''
    if first_text:
        for s in stores:
            s['business_name'] = first_text

    # Second row might have city/category info
    if len(rows) > 1:
        second_text = clean(rows[1][0]) if rows[1] else ''
        if second_text:
            for s in stores:
                s['description'] = second_text

    return stores


def _parse_hours(rows):
    """Parse hours sheet. Handles location/type/days/open/close format."""
    hours = []
    header = rows[0] if rows else []
    start = 1 if header else 0

    for row in rows[start:]:
        vals = [clean(c) for c in (row or [])]
        if not any(vals): continue
        # Skip sub-headers
        if any(k in vals[0].lower() for k in ['upcoming', '향후', 'schedule']): break

        if len(vals) >= 4:
            # Format: store | type | days | open | close
            if len(vals) >= 5 and vals[3] and ':' in str(vals[3]):
                hours.append({
                    'store': vals[0], 'type': vals[1], 'days': vals[2],
                    'open': vals[3], 'close': vals[4] if len(vals) > 4 else ''
                })
            # Format: type | days | open | close (no store column)
            elif vals[1] and ':' in str(vals[2]):
                hours.append({
                    'store': '', 'type': vals[0], 'days': vals[1],
                    'open': vals[2], 'close': vals[3]
                })
    return hours


def _parse_menu_sheet(rows):
    """Parse a menu sheet with category headers and items."""
    items = []
    current_cat = 'Uncategorized'
    # Skip header row
    start = 0
    if rows:
        h = clean(rows[0][0]).lower()
        if h in ('no', 'no.', '#', 'number', '번호', 'item'):
            start = 1

    for row in rows[start:]:
        vals = list(row or []) + [None]*10
        a = clean(vals[0])
        b = clean(vals[1])
        c = clean(vals[2])  # price
        d = clean(vals[3])  # description
        e = clean(vals[4])  # option groups
        f = clean(vals[5])  # option details

        if not a and not b: continue

        # Category header: emoji prefix or non-numeric A with no price
        if a and not a.replace('.','').isdigit():
            has_emoji = bool(re.search(r'[\U00010000-\U0010FFFF\u2600-\u27BF]', a))
            if has_emoji or (not c and not b):
                cat_name = re.sub(r'^[\U00010000-\U0010FFFF\u2600-\u27BF\s📌]+', '', a).strip()
                if cat_name: current_cat = cat_name
                continue

        # Menu item: numeric A with name in B
        if a.replace('.','').isdigit() and b:
            items.append({
                'category': current_cat,
                'name': b,
                'price': parse_price(c),
                'description': d if d.lower() not in ('none','') else '',
                'option_groups': e if e.lower() not in ('none','') else '',
                'options_text': f if f.lower() not in ('none','') else '',
            })
        # Fallback: A is name, B is price
        elif a and not a.replace('.','').isdigit() and parse_price(b) > 0:
            items.append({
                'category': current_cat,
                'name': a,
                'price': parse_price(b),
                'description': c,
                'option_groups': '',
                'options_text': '',
            })

    return items


def _parse_options(rows):
    """Parse options sheet."""
    options = []
    start = 0
    if rows:
        h = ' '.join(clean(c) for c in (rows[0] or []))
        if 'option' in h.lower() or '옵션' in h.lower() or '메뉴' in h.lower():
            start = 1

    for row in rows[start:]:
        vals = list(row or []) + [None]*10
        cleaned = [clean(v) for v in vals[:7]]
        if not any(cleaned[:5]): continue

        # Detect format: with or without store column
        # With store: store | item | price | group | option | add_price | type
        # Without:    item | price | group | option | add_price | type
        if len(cleaned) >= 7 and cleaned[0] and not '$' in cleaned[0] and cleaned[2] and '$' in str(vals[2] or ''):
            options.append({
                'store': cleaned[0], 'item': cleaned[1],
                'base_price': parse_price(vals[2]),
                'group': cleaned[3], 'option': cleaned[4],
                'add_price': parse_price(vals[5]),
                'type': cleaned[6] or 'checkbox',
            })
        elif cleaned[0] and cleaned[2]:
            options.append({
                'store': '', 'item': cleaned[0],
                'base_price': parse_price(vals[1]),
                'group': cleaned[2], 'option': cleaned[3],
                'add_price': parse_price(vals[4]),
                'type': cleaned[5] or 'checkbox',
            })
    return options


def _parse_compare(rows):
    items = []
    for row in rows[1:]:
        vals = [clean(c) for c in (row or [])]
        if vals and vals[0] and not vals[0].startswith(' '):
            items.append(vals)
    return items


# ─── Google Maps URL Parser ───────────────────────────────────────────────

def fetch_place_info(query):
    """Extract store info from Google Maps URL or search query."""
    result = {'name':'','address':'','phone':'','hours':[],'rating':'',
              'category':'','website':'','lat':'','lng':''}
    try:
        if 'google.com/maps' in query or 'goo.gl' in query or 'maps.app' in query:
            m = re.search(r'/place/([^/@]+)', query)
            if m:
                from urllib.parse import unquote
                result['name'] = unquote(m.group(1).replace('+', ' '))
            m = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', query)
            if m:
                result['lat'] = m.group(1)
                result['lng'] = m.group(2)
            try:
                resp = http_req.get(query, headers={'User-Agent':'Mozilla/5.0'}, timeout=10, allow_redirects=True)
                text = resp.text
                m = re.search(r'<title>([^<]+)</title>', text)
                if m:
                    title = re.sub(r'\s*[-–]\s*Google.*$', '', m.group(1)).strip()
                    if title: result['name'] = title
                m = re.search(r'"formatted_phone_number":"([^"]+)"', text)
                if m: result['phone'] = m.group(1)
                m = re.search(r'"formatted_address":"([^"]+)"', text)
                if m: result['address'] = m.group(1)
            except: pass
        else:
            result['name'] = query
    except: pass
    return result


# ─── CARTE POS Excel Generator ────────────────────────────────────────────

# Styles
HF = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HFL = PatternFill(start_color='006EFF', end_color='006EFF', fill_type='solid')
SF = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
TB = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
            top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
WA = Alignment(vertical='center', wrap_text=True)

def _styled_sheet(ws, headers, rows, widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HF; c.fill = HFL; c.alignment = CA; c.border = TB
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Calibri', size=10)
            c.alignment = WA; c.border = TB
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'


def generate_carte_excel(parsed, store_info_override=None, settings=None):
    """Generate the complete CARTE POS setup Excel from parsed data."""
    wb = Workbook()
    settings = settings or {}

    # Pick first store or use override
    stores = parsed.get('stores', [])
    store = stores[0] if stores else {}
    if store_info_override:
        store.update({k:v for k,v in store_info_override.items() if v})
    biz_name = store.get('business_name', store.get('name', 'Store'))

    # ── Sheet 1: Store Info ──
    ws1 = wb.active; ws1.title = 'Store Info'
    addr = store.get('address', '')
    # Try to parse US address
    city = state = zipcode = street = ''
    if addr:
        m = re.match(r'(.+?),\s*([^,]+),\s*([A-Z]{2})\s*(\d{5})?', addr)
        if m:
            street = m.group(1); city = m.group(2); state = m.group(3); zipcode = m.group(4) or ''
        else:
            street = addr

    info_rows = [
        ['Business Name', biz_name],
        ['Store Name', store.get('name', biz_name)],
        ['Phone', store.get('phone', '')],
        ['Email', store.get('email', '')],
        ['Country', 'US'],
        ['State', state or store.get('state', '')],
        ['City', city or store.get('city', '')],
        ['ZIP Code', zipcode or store.get('zip', '')],
        ['Street', street or store.get('street', '')],
        ['Suite/Unit', store.get('suite', '')],
        ['Timezone', store.get('timezone', 'America/Los_Angeles')],
        ['Website', store.get('website', store.get('online_url', ''))],
        ['Latitude', store.get('lat', '')],
        ['Longitude', store.get('lng', '')],
    ]
    _styled_sheet(ws1, ['Field', 'Value'], info_rows, [22, 50])

    # ── Sheet 2: Menu Import (CARTE POS format) ──
    ws2 = wb.create_sheet('Menu Import')
    all_items = []
    for store_name, items in parsed.get('menus', {}).items():
        for it in items:
            allergen_info = it.get('allergen_info', '')
            if not allergen_info:
                # Build from lists if available
                parts = it.get('allergens', []) + it.get('dietary', [])
                allergen_info = ', '.join(parts) if parts else ''
            all_items.append([
                it.get('category', ''),
                it.get('name', ''),
                it.get('description', ''),
                it.get('price', 0),
                '',  # item_cost
                allergen_info,
                'Y', # included_tax
                'Y', # visible
                it.get('image_url', ''),
            ])
    _styled_sheet(ws2,
        ['item_category', 'item_name', 'description', 'price', 'item_cost', 'allergen_info', 'included_tax', 'visible', 'image_url'],
        all_items,
        [20, 30, 40, 10, 10, 20, 14, 10, 40])

    # ── Sheet 3: Menu Full (with options info) ──
    ws3 = wb.create_sheet('Menu Detail')
    detail_rows = []
    for store_name, items in parsed.get('menus', {}).items():
        for it in items:
            detail_rows.append([
                store_name, it.get('category',''), it.get('name',''),
                it.get('price',0), it.get('description',''),
                it.get('option_groups',''), it.get('options_text',''),
            ])
    _styled_sheet(ws3,
        ['Store', 'Category', 'Item', 'Price', 'Description', 'Option Groups', 'Option Details'],
        detail_rows,
        [18, 16, 28, 10, 35, 20, 45])

    # ── Sheet 4: Business Hours ──
    ws4 = wb.create_sheet('Business Hours')
    hours = parsed.get('hours', [])
    if hours:
        hour_rows = [[h.get('store',''), h.get('type',''), h.get('days',''),
                       h.get('open',''), h.get('close','')] for h in hours]
    else:
        # Default hours
        hour_rows = [
            ['', 'Dine-in', 'Mon - Fri', '10:00', '21:00'],
            ['', 'Dine-in', 'Sat', '11:00', '20:00'],
            ['', 'Dine-in', 'Sun', '11:00', '20:00'],
            ['', 'Pickup', 'Mon - Fri', '10:00', '21:00'],
            ['', 'Pickup', 'Sat', '11:00', '20:00'],
            ['', 'Delivery', 'Mon - Fri', '10:00', '20:00'],
        ]
    _styled_sheet(ws4, ['Store', 'Type', 'Days', 'Open', 'Close'], hour_rows,
                  [20, 14, 16, 10, 10])

    # ── Sheet 5: Options ──
    ws5 = wb.create_sheet('Options')
    opt_rows = []
    # Deduplicate options by group+option
    seen_opts = {}
    for o in parsed.get('options', []):
        key = f"{o.get('group','')}|{o.get('option','')}"
        if key not in seen_opts:
            seen_opts[key] = o
            opt_rows.append([o.get('group',''), o.get('option',''),
                           o.get('add_price',0), o.get('type','checkbox')])
    _styled_sheet(ws5, ['Option Group', 'Option Name', 'Additional Price', 'Selection Type'],
                  opt_rows, [22, 25, 16, 14])

    # ── Sheet 6: Settings ──
    ws6 = wb.create_sheet('Settings')
    order_types = settings.get('order_types', ['Dine In', 'Take Out', 'Pickup', 'Delivery'])
    payment_methods = settings.get('payment_methods', ['Cash', 'Credit Card', 'Debit Card', 'Apple Pay', 'Google Pay'])
    tips = settings.get('tips', [15, 18, 20, 25])
    tax_rate = settings.get('tax_rate', 10.25)

    setting_rows = [
        ['--- Order Types ---', ''],
    ] + [[ot, 'Y'] for ot in order_types] + [
        ['', ''],
        ['--- Payment Methods ---', ''],
    ] + [[pm, 'Y'] for pm in payment_methods] + [
        ['', ''],
        ['--- Tax ---', ''],
        ['Sales Tax Rate (%)', tax_rate],
        ['Tax Type', 'Exclusive'],
        ['', ''],
        ['--- Tips ---', ''],
    ] + [[f'Tip Preset {i+1} (%)', t] for i, t in enumerate(tips)]
    _styled_sheet(ws6, ['Setting', 'Value'], setting_rows, [28, 20])
    # Style section headers
    for ri in range(2, len(setting_rows)+2):
        cell = ws6.cell(row=ri, column=1)
        if cell.value and str(cell.value).startswith('---'):
            cell.value = str(cell.value).replace('---','').strip()
            cell.font = Font(name='Calibri', bold=True, size=11, color='006EFF')
            cell.fill = SF

    # ── Sheet 7: Category Summary ──
    ws7 = wb.create_sheet('Category Summary')
    cat_rows = []
    for store_name, items in parsed.get('menus', {}).items():
        cats = {}
        for it in items:
            cat = it.get('category', 'Uncategorized')
            cats[cat] = cats.get(cat, 0) + 1
        for cat, cnt in cats.items():
            cat_rows.append([store_name, cat, cnt])
    _styled_sheet(ws7, ['Store', 'Category', 'Item Count'], cat_rows, [20, 25, 12])

    # ── Sheet 8: Multi-Store Info (if multiple) ──
    if len(stores) > 1:
        ws8 = wb.create_sheet('All Locations')
        loc_rows = []
        for s in stores:
            loc_rows.append([s.get('name',''), s.get('address',''), s.get('phone',''),
                           s.get('lat',''), s.get('lng',''), s.get('website','')])
        _styled_sheet(ws8, ['Location', 'Address', 'Phone', 'Lat', 'Lng', 'Website'],
                     loc_rows, [22, 40, 18, 14, 14, 30])

    # Save
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    slug = re.sub(r'[^\w]', '_', biz_name)[:30]
    filename = f'CARTE_Setup_{slug}_{ts}.xlsx'
    filepath = os.path.join(UPLOAD_DIR, filename)
    wb.save(filepath)
    return filepath, filename


def generate_menu_import_excel(parsed, store_filter=None):
    """Generate just the CARTE POS menu import file."""
    wb = Workbook()
    ws = wb.active; ws.title = 'Menu Import'
    items = []
    for store_name, menu in parsed.get('menus', {}).items():
        if store_filter and store_filter != store_name: continue
        for it in menu:
            allergen_info = it.get('allergen_info', '')
            if not allergen_info:
                parts = it.get('allergens', []) + it.get('dietary', [])
                allergen_info = ', '.join(parts) if parts else ''
            items.append([it.get('category',''), it.get('name',''),
                         it.get('description',''), it.get('price',0),
                         '', allergen_info, 'Y', 'Y',
                         it.get('image_url', '')])
    _styled_sheet(ws,
        ['item_category','item_name','description','price','item_cost','allergen_info','included_tax','visible','image_url'],
        items, [20, 30, 40, 10, 10, 20, 14, 10, 40])

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'CARTE_Menu_Import_{ts}.xlsx'
    filepath = os.path.join(UPLOAD_DIR, filename)
    wb.save(filepath)
    return filepath, filename


# ─── Routes ────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/parse-menu', methods=['POST'])
def api_parse_menu():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Empty filename'}), 400

    safe_name = re.sub(r'[^\w.\-]', '_', f.filename)
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    f.save(filepath)

    try:
        parsed = parse_uploaded_excel(filepath)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 400

    # Flatten for frontend
    all_items = []
    for store_name, items in parsed.get('menus', {}).items():
        for it in items:
            it['store'] = store_name
            all_items.append(it)

    # Category summary
    cat_summary = {}
    for it in all_items:
        cat = it.get('category', '?')
        cat_summary[cat] = cat_summary.get(cat, 0) + 1

    return jsonify({
        'success': True,
        'filename': f.filename,
        'stores': parsed.get('stores', []),
        'hours': parsed.get('hours', []),
        'items': all_items,
        'options': parsed.get('options', [])[:500],  # limit for preview
        'total_options': len(parsed.get('options', [])),
        'category_summary': cat_summary,
        'sheets_found': parsed.get('sheets_found', []),
        'total_items': len(all_items),
    })


@app.route('/api/fetch-bentobox', methods=['POST'])
def api_fetch_bentobox():
    """Fetch menu data from a BentoBox restaurant website."""
    data = request.get_json(silent=True) or {}
    url = data.get('url', '').strip()
    if not url:
        return jsonify({'error': 'No URL provided'}), 400
    if not is_bentobox_url(url):
        return jsonify({'error': 'Not a valid BentoBox URL (must be *.getbento.com)'}), 400
    try:
        result = fetch_bentobox_data(url)
        return jsonify(result)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'BentoBox fetch failed: {e}'}), 500


@app.route('/api/fetch-place', methods=['POST'])
def api_fetch_place():
    data = request.get_json(silent=True) or {}
    query = data.get('query', '').strip()
    if not query:
        return jsonify({'error': 'No query provided'}), 400

    # Detect BentoBox URLs and redirect to BentoBox fetcher
    if is_bentobox_url(query):
        try:
            result = fetch_bentobox_data(query)
            return jsonify(result)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            return jsonify({'error': f'BentoBox fetch failed: {e}'}), 500

    # Google Maps or generic query
    result = fetch_place_info(query)
    return jsonify({'success': True, 'result': result})


@app.route('/api/generate-excel', methods=['POST'])
def api_generate_excel():
    data = request.get_json(silent=True) or {}

    # Check if we have BentoBox data
    bentobox_data = data.get('bentobox_data')
    upload_file = data.get('upload_filename', '')
    store_info = data.get('store_info', {})
    settings = data.get('settings', {})

    if bentobox_data:
        # Convert BentoBox data to parsed format
        parsed = bentobox_to_parsed(bentobox_data)
    elif upload_file:
        filepath = os.path.join(UPLOAD_DIR, re.sub(r'[^\w.\-]', '_', upload_file))
        if os.path.exists(filepath):
            parsed = parse_uploaded_excel(filepath)
        else:
            parsed = {'menus': {}, 'stores': [], 'hours': [], 'options': []}
    else:
        # Build parsed from submitted items
        items = data.get('items', [])
        parsed = {
            'menus': {},
            'stores': [store_info] if store_info else [],
            'hours': data.get('hours', []),
            'options': data.get('options', []),
        }
        # Group items by store
        for it in items:
            store = it.get('store', 'Default')
            if store not in parsed['menus']:
                parsed['menus'][store] = []
            parsed['menus'][store].append(it)

    filepath, filename = generate_carte_excel(parsed, store_info, settings)
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/generate-menu-excel', methods=['POST'])
def api_generate_menu_excel():
    data = request.get_json(silent=True) or {}
    items = data.get('items', []) or data.get('menu_items', [])
    store_filter = data.get('store_filter', None)

    parsed = {'menus': {}}
    for it in items:
        store = it.get('store', 'Default')
        if store not in parsed['menus']:
            parsed['menus'][store] = []
        parsed['menus'][store].append(it)

    filepath, filename = generate_menu_import_excel(parsed, store_filter)
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/push-to-pos', methods=['POST'])
def api_push_to_pos():
    data = request.get_json(silent=True) or {}
    login_id = data.get('login_id', '')
    password = data.get('password', '')
    items = data.get('items', [])
    store_info = data.get('store_info', {})
    settings = data.get('settings', {})

    if not login_id or not password:
        return jsonify({'error': 'Login credentials required'}), 400

    results = []
    errors = []

    # Login
    try:
        resp = http_req.post(f'{CARTE_API}/api/users/login',
            json={'loginId': login_id, 'password': password}, timeout=15)
        if resp.status_code != 200:
            return jsonify({'error': 'Login failed'}), 401
        rd = resp.json().get('resultData', {})
        token = rd.get('token', '')
        branch_id = rd.get('branchId', '')
        if not token:
            return jsonify({'error': 'No token received'}), 401
        results.append({'step': 'Login', 'status': 'success', 'branchId': branch_id})
    except Exception as e:
        return jsonify({'error': f'Login failed: {e}'}), 500

    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}

    # Create categories then import menu
    if items:
        categories = list(set(it.get('category','') for it in items if it.get('category')))
        for cat in categories:
            try:
                r = http_req.post(f'{CARTE_API}/api/menu/category/save',
                    json={'branchId': branch_id, 'categoryName': cat}, headers=headers, timeout=15)
                results.append({'step': f'Category: {cat}', 'status': 'success' if r.status_code==200 else 'failed'})
            except Exception as e:
                errors.append({'step': f'Category: {cat}', 'error': str(e)})

        # Import in batches
        batch = []
        for it in items:
            allergen_info = it.get('allergen_info', '')
            if not allergen_info:
                parts = it.get('allergens', []) + it.get('dietary', [])
                allergen_info = ', '.join(parts) if parts else ''
            item_payload = {
                'item_category': it.get('category',''), 'item_name': it.get('name',''),
                'description': it.get('description',''), 'price': it.get('price',0),
                'item_cost': 0, 'allergen_info': allergen_info,
                'included_tax': 'Y', 'visible': 'Y',
            }
            if it.get('image_url'):
                item_payload['image_url'] = it['image_url']
            batch.append(item_payload)
            if len(batch) >= 50:
                try:
                    r = http_req.post(f'{CARTE_API}/api/menu/import',
                        json={'branchId': branch_id, 'items': batch}, headers=headers, timeout=30)
                    results.append({'step': f'Import {len(batch)} items', 'status': 'success' if r.status_code==200 else 'failed'})
                except Exception as e:
                    errors.append({'step': 'Menu import', 'error': str(e)})
                batch = []
        if batch:
            try:
                r = http_req.post(f'{CARTE_API}/api/menu/import',
                    json={'branchId': branch_id, 'items': batch}, headers=headers, timeout=30)
                results.append({'step': f'Import {len(batch)} items', 'status': 'success' if r.status_code==200 else 'failed'})
            except Exception as e:
                errors.append({'step': 'Menu import', 'error': str(e)})

    return jsonify({'success': len(errors)==0, 'results': results, 'errors': errors, 'branch_id': branch_id})


if __name__ == '__main__':
    print('=' * 50)
    print('  CARTE POS Setup Tool')
    print('  http://localhost:5050')
    print('=' * 50)
    app.run(debug=True, host='0.0.0.0', port=5050)
